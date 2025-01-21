import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image as PILImage, ImageTk

def format_date(date_value):
    #Convert a date to the dd-mm-yyyy format, stripping any time part
    if isinstance(date_value, str):
        # Use regex to strip time part if it exists
        date_value = re.sub(r"\s+\d{2}:\d{2}:\d{2}$", "", date_value.strip())
        try:
            date_value = pd.to_datetime(date_value).strftime('%d-%m-%Y')
        except Exception:
            return date_value  # Return the original if parsing fails
    elif isinstance(date_value, pd.Timestamp):
        return date_value.strftime('%d-%m-%Y')
    elif isinstance(date_value, datetime):
        return date_value.strftime('%d-%m-%Y')
    return date_value

def find_excel_file(folder_path, prefix):
    #Find the first Excel file in the folder with the given prefix
    for file_name in os.listdir(folder_path):
        if file_name.startswith(prefix) and file_name.endswith('.xlsx'):
            return os.path.join(folder_path, file_name)
    return None

def extract_tables_from_holding(file_path):
    # Extract the two tables from the Holding Report
    df = pd.read_excel(file_path, sheet_name=0, header=None)

    # Extracting First table
    start_row_1 = df[df.iloc[:, 0] == 'Client Equity Code/UCID/Name'].index[0]
    end_row_1 = start_row_1 + 4  # First table has 4 rows
    first_table = df.iloc[start_row_1:end_row_1, :2]
    first_table.columns = ['Field', 'Value']
    
    # Format dates in the first table
    first_table['Value'] = first_table['Value'].apply(format_date)

    # Extracing Second table (dynamic location based on headers)
    second_table_start = df[df.iloc[:, 0] == 'Instrument Name'].index[0]
    second_table = pd.read_excel(file_path, sheet_name=0, header=second_table_start)

    # Filter columns to exclude the unwanted ones
    columns_to_remove = [
        'ST Qty', 'ST G/L', 'LT Qty', 'LT G/L',
        'UnrealisedGain/Loss%', 'ISIN'
    ]
    second_table = second_table.drop(columns=columns_to_remove, errors='ignore')

    # Filter remaining required columns
    required_columns = [
        'Instrument Name', 'Quantity', 'Purchase Price', 'Purchase Value',
        'Market Price', 'Market Value', 'UnrealisedGain/Loss'
    ]
    second_table = second_table[required_columns]

    return first_table, second_table


def process_export_file(file_path):
    # Processing the Export file to calculate values for the report
    df = pd.read_excel(file_path)
    df['CREDIT'] = pd.to_numeric(df['CREDIT'].replace({',': ''}, regex=True), errors='coerce')
    df['DEBIT'] = pd.to_numeric(df['DEBIT'].replace({',': ''}, regex=True), errors='coerce')

    # Pay-In Details
    pay_in_details = df[df['VOUCHER TYPE'] == 'PAYIN'][['VOUCHER DATE', 'EFFECTIVE DATE', 'VOUCHER TYPE', 'CREDIT']]
    pay_in_details.columns = ['Voucher Date', 'Effective Date', 'Voucher Type', 'Amount']
    pay_in_details['Voucher Date'] = pay_in_details['Voucher Date'].apply(format_date)
    pay_in_details['Effective Date'] = pay_in_details['Effective Date'].apply(format_date)
    pay_in_total = pay_in_details['Amount'].sum()
    pay_in_details = pay_in_details.append({
        'Voucher Date': 'Total', 'Effective Date': '', 'Voucher Type': '', 'Amount': pay_in_total
    }, ignore_index=True)

    # Pay-Out Details
    pay_out_details = df[df['VOUCHER TYPE'] == 'PAYOUT'][['VOUCHER DATE', 'EFFECTIVE DATE', 'VOUCHER TYPE', 'DEBIT']]
    pay_out_details.columns = ['Voucher Date', 'Effective Date', 'Voucher Type', 'Amount']
    pay_out_details['Voucher Date'] = pay_out_details['Voucher Date'].apply(format_date)
    pay_out_details['Effective Date'] = pay_out_details['Effective Date'].apply(format_date)
    pay_out_total = pay_out_details['Amount'].sum()
    pay_out_details = pay_out_details.append({
        'Voucher Date': 'Total', 'Effective Date': '', 'Voucher Type': '', 'Amount': pay_out_total
    }, ignore_index=True)

    # Initial Investment
    initial_investment_date = format_date(df['EFFECTIVE DATE'].iloc[-1])
    initial_investment_amount = df['CREDIT'].iloc[-1]
    # Additional Investment
    additional_investment = df[df['VOUCHER TYPE'] == 'PAYIN']['CREDIT'].sum() - initial_investment_amount

    # Amount Paid back
    amount_paid_back = df[df['VOUCHER TYPE'] == 'PAYOUT']['DEBIT'].sum()

    return initial_investment_date, initial_investment_amount, additional_investment, amount_paid_back, pay_in_details, pay_out_details

def process_dividend_file(file_path):
    # Processing the Dividend file to get Dividend Received
    df = pd.read_excel(file_path)
    dividend_received = pd.to_numeric(df.iloc[-1, 3], errors='coerce')
    return dividend_received


def generate_investment_report(folder_path, output_folder):
    # Generate the investment report based on the provided files
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    holding_file = find_excel_file(folder_path, 'Holding')
    export_file = find_excel_file(folder_path, 'Export')
    dividend_file = find_excel_file(folder_path, 'Dividend')

    if not (holding_file and export_file and dividend_file):
        raise FileNotFoundError("One or more required files are missing in the folder.")

    # Extracting data from Holding file
    holding_table_1, holding_table_2 = extract_tables_from_holding(holding_file)
    
    client_equity_value = holding_table_1[holding_table_1['Field'] == 'Client Equity Code/UCID/Name']['Value'].values[0]
    cleaned_client_equity_value = client_equity_value.replace('/', '-')  # Replace '/' with '-'
    
    output_file_name = f"Generated Report - {cleaned_client_equity_value}.xlsx"
    output_path = os.path.join(output_folder, output_file_name)

    # Processing Export file
    (
        initial_investment_date,
        initial_investment_amount,
        additional_investment,
        amount_paid_back,
        pay_in_details,
        pay_out_details,
    ) = process_export_file(export_file)

    # Processing Dividend file
    dividend_received = process_dividend_file(dividend_file)

    todays_value_date = holding_table_1[holding_table_1['Field'] == 'Report Generation Date']['Value'].values[0]
    todays_value = pd.to_numeric(holding_table_2['Market Value'].iloc[-1], errors='coerce')

    # Converting extracted values to numeric
    initial_investment_amount = float(initial_investment_amount)
    additional_investment = float(additional_investment)
    amount_paid_back = float(amount_paid_back)
    todays_value = float(todays_value)
    dividend_received = float(dividend_received)

    # calculations
    total_investment = initial_investment_amount + additional_investment
    net_investment = total_investment - amount_paid_back
    net_profit = todays_value - net_investment
    total_profit = net_profit + dividend_received
    absolute_profit_percentage = (total_profit / net_investment) * 100

    # Creating the second table
    third_table = pd.DataFrame({
        'Particulars': [
            'Initial Investment', 'Additional Investment', 'Total Investment',
            'Amount Paid back', 'Net Investment', 'Todays Value', 'Net Profit',
            'Dividend Received', 'Total Profit', 'Absolute Profit %'
        ],
        'Date': [
            initial_investment_date, '-', '-', '-', '-', todays_value_date, '-', '-', '-', '-'
        ],
        'Amount (Rs.)': [
            initial_investment_amount, additional_investment, total_investment,
            amount_paid_back, net_investment, todays_value, net_profit,
            dividend_received, total_profit, absolute_profit_percentage
        ] 
    })

    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Writing the first table (Holding Table 1)
        holding_table_1.to_excel(writer, sheet_name='Report', index=False, startrow=0)

        # Writing the second table
        start_row_2 = len(holding_table_1) + 3  # Adding three empty rows between the tables
        third_table.to_excel(writer, sheet_name='Report', index=False, startrow=start_row_2)

        # Writing the third table
        start_row_3 = start_row_2 + len(third_table) + 3
        holding_table_2.to_excel(writer, sheet_name='Report', index=False, startrow=start_row_3)

        # Writing Pay-In Details table
        start_row_4 = start_row_3 + len(holding_table_2) + 3
        pay_in_details.to_excel(writer, sheet_name='Report', index=False, startrow=start_row_4)

        # Writing Pay-Out Details table
        start_row_5 = start_row_4 + len(pay_in_details) + 3
        pay_out_details.to_excel(writer, sheet_name='Report', index=False, startrow=start_row_5)

    
    wb = load_workbook(output_path)
    ws = wb['Report']
    ws.sheet_view.showGridLines = False

    # Defining header and border styles
    header_fill = PatternFill(start_color='E8A820', end_color='E8A820', fill_type='solid')
    header_font = Font(bold=True, color='000772')
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Formatting the first table (from holding report)
    ws.merged_cells.ranges.add("A8:B8")
    ws.cell(row=1, column=1, value='Holding-As On Date').fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    for row in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
        
    last_row_2 = start_row_3 + len(holding_table_2) + 1
    for cell in ws.iter_rows(min_row=last_row_2, max_row=last_row_2, min_col=1, max_col=7):
        for c in cell:
            c.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            c.font = Font(bold=True, color="e05904")
 
    # Formatting the last row of Pay-In Details
    last_row_4 = start_row_4 + len(pay_in_details) + 1
    for cell in ws.iter_rows(min_row=last_row_4, max_row=last_row_4, min_col=1, max_col=4):
        for c in cell:
            c.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            c.font = Font(bold=True, color="e05904")

    # Formatting the last row of Pay-Out Details
    last_row_5 = start_row_5 + len(pay_out_details) + 1
    for cell in ws.iter_rows(min_row=last_row_5, max_row=last_row_5, min_col=1, max_col=4):
        for c in cell:
            c.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            c.font = Font(bold=True, color="e05904")
    
    # Formatting the second table (Generated Table)
    header_row_2 = start_row_2 + 1
    for cell in ws.iter_rows(min_row=header_row_2, max_row=header_row_2, min_col=1, max_col=3):
        for c in cell:
            c.fill = header_fill
            c.font = header_font
    for row in ws.iter_rows(min_row=start_row_2+1, max_row=start_row_2+len(third_table)+1, min_col=1, max_col=3):
        for cell in row:
            cell.border = border

    # Formatting the third table (from holding report)
    header_row_3 = start_row_3 + 1
    for cell in ws.iter_rows(min_row=header_row_3, max_row=header_row_3, min_col=1, max_col=7):
        for c in cell:
            c.fill = header_fill
            c.font = header_font
    for row in ws.iter_rows(min_row=start_row_3+1, max_row=start_row_3+len(holding_table_2)+1, min_col=1, max_col=7):
        for cell in row:
            cell.border = border

    # Formatting the Pay-In Details table
    header_row_4 = start_row_4 + 1
    for cell in ws.iter_rows(min_row=header_row_4, max_row=header_row_4, min_col=1, max_col=4):
        for c in cell:
            c.fill = header_fill
            c.font = header_font
    for row in ws.iter_rows(min_row=start_row_4+1, max_row=start_row_4+len(pay_in_details)+1, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    # Formatting the Pay-Out Details table
    header_row_5 = start_row_5 + 1
    # Formatting headers specifically for 4 columns
    for cell in ws.iter_rows(min_row=header_row_5, max_row=header_row_5, min_col=1, max_col=4):
        for c in cell:
            c.fill = header_fill
            c.font = header_font
    for row in ws.iter_rows(min_row=start_row_5+1, max_row=start_row_5+len(pay_out_details)+1, min_col=1, max_col=4):
        for cell in row:
            cell.border = border

    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws.merged_cells.ranges.add("A1:P6")

    
    logo_path = 'img/logo for letterhead.png' 
    img = Image(logo_path)
    img.width = 990 
    img.height = 120
    ws.add_image(img, 'A1')

    # Save the workbook
    wb.save(output_path)

def browse_input_folder():
    folder_selected = filedialog.askdirectory(title="Select Input Folder")
    input_folder_var.set(folder_selected)

def browse_output_folder():
    folder_selected = filedialog.askdirectory(title="Select Output Folder")
    output_folder_var.set(folder_selected)

def generate_report():
    try:
        input_folder = input_folder_var.get()
        output_folder = output_folder_var.get()
        
        if not input_folder or not output_folder:
            raise ValueError("Both input and output folders must be selected.")
        
        report_path = generate_investment_report(input_folder, output_folder)
        messagebox.showinfo("Success", f"Report generated successfully!")
    except Exception as e:
        messagebox.showerror("Error", str(e))

root = tk.Tk()
root.title("SPB Finserv Portfolio Report Generator")

logo_path = 'img/logo.png'
if os.path.exists(logo_path):
    logo_image = PILImage.open(logo_path)
    logo_image = logo_image.resize((300, 110), PILImage.Resampling.LANCZOS)  # Resize if necessary
    logo_photo = ImageTk.PhotoImage(logo_image)
    logo_label = tk.Label(root, image=logo_photo)
    logo_label.image = logo_photo  # Keep a reference to the image
    logo_label.grid(row=0, column=0, columnspan=3, pady=10)


input_folder_var = tk.StringVar()
output_folder_var = tk.StringVar()

input_label = tk.Label(root, text="Select Input Folder:")
input_button = tk.Button(root, text="Browse", command=browse_input_folder)
input_entry = tk.Entry(root, textvariable=input_folder_var, width=50)

output_label = tk.Label(root, text="Select Output Folder:")
output_button = tk.Button(root, text="Browse", command=browse_output_folder)
output_entry = tk.Entry(root, textvariable=output_folder_var, width=50)

generate_button = tk.Button(root, text="Generate Report", command=generate_report)

input_label.grid(row=1, column=0, padx=10, pady=10)
input_button.grid(row=1, column=1, padx=10, pady=10)
input_entry.grid(row=1, column=2, padx=10, pady=10)

output_label.grid(row=2, column=0, padx=10, pady=10)
output_button.grid(row=2, column=1, padx=10, pady=10)
output_entry.grid(row=2, column=2, padx=10, pady=10)

generate_button.grid(row=3, column=0, columnspan=3, pady=20)

# Run the application
root.mainloop()